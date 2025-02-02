import openpyxl
import streamlit as st
import pandas as pd
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from copy import copy
import os
import tempfile
from PIL import Image as PILImage  # PillowのImageクラスをインポート
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break


# タイトルをページの最上部に配置
#st.title('🛠️パンフレット作成🛠️')
st.markdown(
        "<h3 style='font-size:30px;'>🛠️パンフレット作成🛠️</h3>",  # 'font-size'でサイズを指定
        unsafe_allow_html=True
    )

# ファイルパスを指定してExcelファイルを読み込む
@st.cache_data
def load_data(file_path):
    #読み取り
    df = pd.read_excel(file_path)
    # '肥料名称' カラムから NaN を取り除く
    df = df.dropna(subset=['肥料名称'])
    return df

# キャッシュとセッションステートのクリア
if st.button('cash Clear'):
    # キャッシュをクリア
    st.cache_data.clear()
    st.cache_resource.clear()
    st.rerun()  # アプリをリロード

df = load_data('銘柄データ_BB.xlsx')
df_ekihi = load_data('銘柄データ_液肥.xlsx')
df_kasei = load_data('銘柄データ_化成.xlsx')

# 肥料名称のリストを作る
fertilizer_names = df['肥料名称'].tolist()
fertilizer_names_ekihi = df_ekihi['肥料名称'].tolist()
fertilizer_names_kasei = df_kasei['肥料名称'].tolist()

# 選択されたアイテムのリストを作成
selected_fertilizer = []
selected_fertilizer_ekihi = []
selected_fertilizer_kasei = []

st.markdown(
    """
    <style>
    .main .block-container {
        max-width: 1000px;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

# 3つのカラムを作成
col1, col2, col3 = st.columns(3)

# チェックボックスの状態をセッションステートに保存
if 'selected_fertilizer_bb' not in st.session_state:
    st.session_state.selected_fertilizer_bb = [False] * len(fertilizer_names)
if 'selected_fertilizer_kasei' not in st.session_state:
    st.session_state.selected_fertilizer_kasei = [False] * len(fertilizer_names_kasei)
if 'selected_fertilizer_ekihi' not in st.session_state:
    st.session_state.selected_fertilizer_ekihi = [False] * len(fertilizer_names_ekihi)

# リセットボタンを表示
if st.button('チェックマークをリセット<注意:最後のチェックしたのは消えない>'):
    st.session_state.selected_fertilizer_bb = [False] * len(fertilizer_names)
    st.session_state.selected_fertilizer_kasei = [False] * len(fertilizer_names_kasei)
    st.session_state.selected_fertilizer_ekihi = [False] * len(fertilizer_names_ekihi)
    st.rerun()  # リセット後に再描画

# 1列目に球技のチェックボックスを作成
with col1:
    #st.header("BB")
    # ヘッダーの文字サイズを小さくする
    st.markdown(
        "<h3 style='font-size:25px;'>BB</h3>",  # 'font-size'でサイズを指定
        unsafe_allow_html=True
    )
    for i, fertilizer_name in enumerate(fertilizer_names):
        checkbox_value = st.session_state.selected_fertilizer_bb[i]
        if st.checkbox(fertilizer_name, key=fertilizer_name, value=checkbox_value):
            selected_fertilizer.append(fertilizer_name)
            st.session_state.selected_fertilizer_bb[i] = True
        else:
            st.session_state.selected_fertilizer_bb[i] = False

# 2列目に球技のチェックボックスを作成
with col2:
    ##st.header("化成")
    st.markdown(
        "<h3 style='font-size:25px;'>化成</h3>",  # 'font-size'でサイズを指定
        unsafe_allow_html=True
    )
    for i, fertilizer_name_kasei in enumerate(fertilizer_names_kasei):
        checkbox_value = st.session_state.selected_fertilizer_kasei[i]
        if st.checkbox(fertilizer_name_kasei, key=fertilizer_name_kasei, value=checkbox_value):
            selected_fertilizer_kasei.append(fertilizer_name_kasei)
            st.session_state.selected_fertilizer_kasei[i] = True
        else:
            st.session_state.selected_fertilizer_kasei[i] = False

# 3列目に液肥のチェックボックスを作成
with col3:
    #st.header("液肥")
    st.markdown(
        "<h3 style='font-size:25px;'>液肥</h3>",  # 'font-size'でサイズを指定
        unsafe_allow_html=True
    )
    for i, fertilizer_name_ekihi in enumerate(fertilizer_names_ekihi):
        checkbox_value = st.session_state.selected_fertilizer_ekihi[i]
        if st.checkbox(fertilizer_name_ekihi, key=fertilizer_name_ekihi, value=checkbox_value):
            selected_fertilizer_ekihi.append(fertilizer_name_ekihi)
            st.session_state.selected_fertilizer_ekihi[i] = True
        else:
            st.session_state.selected_fertilizer_ekihi[i] = False
# 選択されたアイテムの数を主翼
selected_fertilizer_count = len(selected_fertilizer)
selected_fertilizer_count_ekihi = len(selected_fertilizer_ekihi)
selected_fertilizer_count_kasei = len(selected_fertilizer_kasei)

#セットアップする
if st.button('セットアップする'):

    if selected_fertilizer_count > 0:
        # ワークブックをロードする
        wb = openpyxl.load_workbook('bb_tem.xlsx')
        # ワークシートを選択する（シート名を指定する）
        ws = wb['BB_テンプレ']

        ## 必要数
        count_number = selected_fertilizer_count  ###ここがチェックされた数字となる。
        #テンプレートを作るところ。
        m = count_number - 1  # ここでチェックをつけられた分だけコピーすることになる。0からカウント
        count = (m // 2)

        for i in range(0, count):
            row_count = 1
            col_count = 14
            col_offset = i * 13

            # コピー元の範囲（例: A1からM44）
            source_range = [[ws.cell(row=r, column=c) for c in range(1, 14)] for r in range(1, 45)]

            # コピー先の左上セル（例: N1）
            dest_start_cell = ws.cell(row=row_count, column=col_count + col_offset)

            def copy_cell(src_cell, dest_cell):
                dest_cell.value = src_cell.value
                if src_cell.has_style:
                    dest_cell.font = copy(src_cell.font)
                    dest_cell.border = copy(src_cell.border)
                    dest_cell.fill = copy(src_cell.fill)
                    dest_cell.number_format = copy(src_cell.number_format)
                    dest_cell.protection = copy(src_cell.protection)
                    dest_cell.alignment = copy(src_cell.alignment)

            # コピー元範囲の行数と列数を取得する
            row_count = len(source_range)
            col_count = len(source_range[0])

            # コピー元範囲をループしてコピー先にペーストする
            for i in range(row_count):
                for j in range(col_count):
                    src_cell = source_range[i][j]
                    dest_cell = ws.cell(row=dest_start_cell.row + i, column=dest_start_cell.column + j)
                    copy_cell(src_cell, dest_cell)

            # 指定された列幅にコピー元とコピー先の列幅を設定する
            specified_widths = [1, 5.67, 8.42, 5.67, 4, 0.84, 8.08, 8.08, 8.08, 8.08, 8.08, 8.08, 8.08]

            # コピー元の列幅を設定する
            for idx, width in enumerate(specified_widths, start=source_range[0][0].column):
                col_letter = openpyxl.utils.get_column_letter(idx)
                ws.column_dimensions[col_letter].width = width

            # コピー先の列幅を設定する
            for idx, width in enumerate(specified_widths, start=dest_start_cell.column):
                col_letter = openpyxl.utils.get_column_letter(idx)
                ws.column_dimensions[col_letter].width = width

        # いらないところを消す
        number = m + 1  # mが0からカウントとなるため、+1とする
        kesu_offset = (number // 2) * 13

        # 奇数の時のみ実行する
        if number % 2 != 0:
            # A1:M44 の範囲のセルをループする
            for row in ws.iter_rows(min_row=24, max_row=42, min_col=1 + kesu_offset, max_col=13 + kesu_offset):
                for cell in row:
                    # セルの文字を消す
                    cell.value = None

                    # セルの罫線を消す
                    cell.border = Border()

                    # セルの背景色を消す (デフォルトは白)
                    cell.fill = PatternFill(fill_type=None)

                    # セルのフォントスタイルをデフォルトにリセット
                    cell.font = Font()

        #ここからデータを入れるところ
        # 選択されたデータ数分入力する。リスト分0スタートなので、+1とする。
        mm = count_number

        i = 0
        # 各肥料名についてループ
        for fertilizer in selected_fertilizer:
            selected_row = df[df['肥料名称'] == fertilizer]
            
        #   for i in range(0, mm):
            row_offset = (i % 2) * 20
            col_offset = (i // 2) * 13
            
            # cはセル番地でH5を取得、NPK、速攻性、被覆尿素まで入れる
            n_base_row = 5
            n_base_column = 8
            
            name = ws.cell(row=n_base_row + row_offset - 1, column=n_base_column + col_offset - 7)
            name.value = selected_row['肥料名称'].values[0]
            
            # N入力
            n = ws.cell(row=n_base_row + row_offset, column=n_base_column + col_offset)
            n.value = selected_row['N'].values[0]
            # P入力
            p = ws.cell(row=n_base_row + row_offset, column=n_base_column + col_offset + 1)
            p.value = selected_row['P'].values[0]
            # K入力
            k = ws.cell(row=n_base_row + row_offset, column=n_base_column + col_offset + 2)
            k.value = selected_row['K'].values[0]
            # 速攻性N
            s_n = ws.cell(row=n_base_row + row_offset, column=n_base_column + col_offset + 3)
            s_n.value = selected_row['速効性'].values[0]
            # 被覆尿素
            h_n = ws.cell(row=n_base_row + row_offset, column=n_base_column + col_offset + 4)
            h_n.value = selected_row['被覆尿素'].values[0]
            # その他
            ano = ws.cell(row=n_base_row + row_offset + 1, column=n_base_column + col_offset)
            ano.value = selected_row['容量②'].values[0]
            # 栽培適正
            #tekisei = ws.cell(row=n_base_row + row_offset + 12, column=n_base_column + col_offset - 6)
            #tekisei.value = selected_row['栽培適正'].values[0]
            # 品種
            hinshu = ws.cell(row=n_base_row + row_offset + 13, column=n_base_column + col_offset - 6)
            hinshu.value = selected_row['品種'].values[0]
            # 特徴①
            tokuchou_1 = ws.cell(row=n_base_row + row_offset + 4, column=n_base_column + col_offset - 1)
            tokuchou_1.value = selected_row['特徴①'].values[0]
            # 特徴②
            tokuchou_2 = ws.cell(row=n_base_row + row_offset + 5, column=n_base_column + col_offset - 1)
            tokuchou_2.value = selected_row['特徴②'].values[0]
            # 特徴③
            tokuchou_3 = ws.cell(row=n_base_row + row_offset + 6, column=n_base_column + col_offset - 1)
            tokuchou_3.value = selected_row['特徴③'].values[0]
            # 特徴④
            tokuchou_4 = ws.cell(row=n_base_row + row_offset + 7, column=n_base_column + col_offset - 1)
            tokuchou_4.value = selected_row['特徴④'].values[0]
            # 特徴⑤
            tokuchou_5 = ws.cell(row=n_base_row + row_offset + 8, column=n_base_column + col_offset - 1)
            tokuchou_5.value = selected_row['特徴⑤'].values[0]
            
            
            # 容器,肥効曲線の画像を貼り付ける
            # スクリプトのディレクトリを取得
            script_dir = os.path.dirname(os.path.abspath(__file__))
            # 絶対パスを生成
            img_path = os.path.join(script_dir, '容器', f'{fertilizer}.jpg')
            img_path2 = os.path.join(script_dir, '肥効曲線', f'{fertilizer}.jpg')

            # ファイルの存在を確認
            if os.path.exists(img_path):
                
                # Pillowで画像を開く
                original_img = PILImage.open(img_path)

                # 画像のリサイズ
                new_size = (190, 257)  # 新しいサイズを指定
                resized_img = original_img.resize(new_size)
            
                # 一時的なファイルを作成
                with tempfile.NamedTemporaryFile(delete=False, suffix=".jpg") as tmp_file:
                    temp_path = tmp_file.name
                    resized_img.save(temp_path)
                    #st.write(f"Image temporarily saved at {temp_path}")

                # openpyxlのImageクラスでリサイズされた画像を読み込む
                img = OpenpyxlImage(temp_path)
                # Excelのセルに画像を貼り付ける位置を指定
                cell_address = ws.cell(row=n_base_row + row_offset + 1, column=n_base_column + col_offset - 6).coordinate
                img.anchor = cell_address

                # 画像をシートに追加
                ws.add_image(img)
                # 一時ファイルを削除
                #os.remove(temp_path) 
            
            else:
                # ファイルが存在しない場合は何もしない
                pass


            if os.path.exists(img_path2):
                
                # Pillowで画像を開く
                original_img2 = PILImage.open(img_path2)

                # 画像のリサイズ
                new_size2 = (440, 170)  # 新しいサイズを指定
                resized_img2 = original_img2.resize(new_size2)
            
                # 一時的なファイルを作成
                with tempfile.NamedTemporaryFile(delete=False, suffix=".jpg") as tmp_file2:
                    temp_path2 = tmp_file2.name
                    resized_img2.save(temp_path2)

                # openpyxlのImageクラスでリサイズされた画像を読み込む
                img2 = OpenpyxlImage(temp_path2)
                # Excelのセルに画像を貼り付ける位置を指定
                cell_address2 = ws.cell(row=n_base_row + row_offset + 10, column=n_base_column + col_offset - 1).coordinate
    
    #            ws.cell(row=n_base_row + row_offset + 1, column=n_base_column + col_offset - 6).coordinate
                img2.anchor = cell_address2

                # 画像をシートに追加
                ws.add_image(img2)
                # 一時ファイルを削除
                #os.remove(temp_path) 
            
            else:
                # ファイルが存在しない場合は何もしない
                pass

            i = i + 1

#=================================
        # 行数と列数を変数として定義
#        from openpyxl.utils import get_column_letter
#        from openpyxl.worksheet.pagebreak import Break
        
        row_start = 1
        row_end = 42
        col_start = 1  # A列は1
        col_end = n_base_column + col_offset + 5

        # 列を文字列に変換
        col_start_letter = get_column_letter(col_start)
        col_end_letter = get_column_letter(col_end)

        # print_areaを設定
        ws.print_area = f'{col_start_letter}{row_start}:{col_end_letter}{row_end}'

        # 13列ごとに垂直改ページを追加
        cols_per_page = 13  # 13列ごとに改ページ
        for col in range(cols_per_page, col_end + 1 , cols_per_page):
            ws.col_breaks.append(Break(id=col))

        # 印刷時の拡大・縮小を85%に設定
        ws.page_setup.scale = 85

#-==================================

        # 変更を保存する
        wb.save('bb_tem_finish.xlsx')


    if selected_fertilizer_count_kasei > 0:
        # ワークブックをロードする
        wb = openpyxl.load_workbook('kasei_tem.xlsx')
        # ワークシートを選択する（シート名を指定する）
        ws = wb['化成_テンプレ']

        # 必要数
        count_number_kasei = selected_fertilizer_count_kasei  ###ここがチェックされた数字となる。
        #テンプレートを作るところ。
        m = count_number_kasei - 1  # ここでチェックをつけられた分だけコピーすることになる。0からカウント
        count_number_kasei = (m // 3)

        for i in range(0, count_number_kasei):
            row_count = 1
            col_count = 14
            col_offset = i * 13
            # コピー元の範囲（例: A1からM50）
            source_range = [[ws.cell(row=r, column=c) for c in range(1, 14)] for r in range(1, 51)]

            # コピー先の左上セル（例: N1）
            dest_start_cell = ws.cell(row=row_count, column=col_count + col_offset)

            def copy_cell(src_cell, dest_cell):
                dest_cell.value = src_cell.value
                if src_cell.has_style:
                    dest_cell.font = copy(src_cell.font)
                    dest_cell.border = copy(src_cell.border)
                    dest_cell.fill = copy(src_cell.fill)
                    dest_cell.number_format = copy(src_cell.number_format)
                    dest_cell.protection = copy(src_cell.protection)
                    dest_cell.alignment = copy(src_cell.alignment)

            # コピー元範囲の行数と列数を取得する
            row_count = len(source_range)
            col_count = len(source_range[0])

            # コピー元範囲をループしてコピー先にペーストする
            for i in range(row_count):
                for j in range(col_count):
                    src_cell = source_range[i][j]
                    dest_cell = ws.cell(row=dest_start_cell.row + i, column=dest_start_cell.column + j)
                    copy_cell(src_cell, dest_cell)

            # 指定された列幅にコピー元とコピー先の列幅を設定する
            specified_widths = [1, 8.42, 5.67, 5.67, 4, 0.84, 8.42, 8.42, 8.42, 8.42, 8.42, 8.42, 8.42]

            # コピー元の列幅を設定する
            for idx, width in enumerate(specified_widths, start=source_range[0][0].column):
                col_letter = openpyxl.utils.get_column_letter(idx)
                ws.column_dimensions[col_letter].width = width

            # コピー先の列幅を設定する
            for idx, width in enumerate(specified_widths, start=dest_start_cell.column):
                col_letter = openpyxl.utils.get_column_letter(idx)
                ws.column_dimensions[col_letter].width = width

    
        # いらないところを消す
        number = m + 1  # mが0からカウントとなるため、+1とする
        kesu_offset = (number // 3) * 13

        # 1,4,7の時のみ実行する
        if (number - 1) % 3 == 0:
            # A1:M44 の範囲のセルをループする
            for row in ws.iter_rows(min_row=20, max_row=50, min_col=1 + kesu_offset, max_col=13 + kesu_offset):
                for cell in row:
                    # セルの文字を消す
                    cell.value = None

                    # セルの罫線を消す
                    cell.border = Border()

                    # セルの背景色を消す (デフォルトは白)
                    cell.fill = PatternFill(fill_type=None)

                    # セルのフォントスタイルをデフォルトにリセット
                    cell.font = Font()

        # 2,5,8の時のみ実行する
        if (number - 2) % 3 == 0:
            # A1:M44 の範囲のセルをループする
            for row in ws.iter_rows(min_row=36, max_row=42, min_col=1 + kesu_offset, max_col=13 + kesu_offset):
                for cell in row:
                    # セルの文字を消す
                    cell.value = None

                    # セルの罫線を消す
                    cell.border = Border()

                    # セルの背景色を消す (デフォルトは白)
                    cell.fill = PatternFill(fill_type=None)

                    # セルのフォントスタイルをデフォルトにリセット
                    cell.font = Font()

        #ここからデータを入れるところ
        #選択されたデータ数分入力する。リスト分0スタートなので、+1とする。
        mm = number

        i = 0
        # 各肥料名についてループ
        for fertilizer in selected_fertilizer_kasei:
            selected_row = df_kasei[df_kasei['肥料名称'] == fertilizer]
            
            # cはセル番地でH5を取得、NPK、速攻性、被覆尿素まで入れる
            n_base_row = 4
            n_base_column = 2

            row_offset = (i % 3) * 16
            col_offset = (i // 3) * 13
            
            # 肥料名称
            name = ws.cell(row=n_base_row + row_offset, column=n_base_column + col_offset)
            name.value = selected_row['肥料名称'].values[0]

            # N
            n = ws.cell(row=n_base_row + row_offset + 1, column=n_base_column + col_offset + 6)
            n.value = selected_row['N'].values[0]

            # P
            n = ws.cell(row=n_base_row + row_offset + 1, column=n_base_column + col_offset + 7)
            n.value = selected_row['P'].values[0]

            # K
            n = ws.cell(row=n_base_row + row_offset + 1, column=n_base_column + col_offset + 8)
            n.value = selected_row['K'].values[0]
            
            # Mg
            n = ws.cell(row=n_base_row + row_offset + 1, column=n_base_column + col_offset + 9)
            n.value = selected_row['Mg'].values[0]

            # Mn
            n = ws.cell(row=n_base_row + row_offset + 1, column=n_base_column + col_offset + 10)
            n.value = selected_row['Mn'].values[0]
            
            # B
            n = ws.cell(row=n_base_row + row_offset + 1, column=n_base_column + col_offset + 11)
            n.value = selected_row['B'].values[0]

            # その他
            for k in range(1,3):
                n = ws.cell(row=n_base_row + row_offset + 1 + k, column=n_base_column + col_offset + 6 )
                n.value = selected_row[f'その他{k}'].values[0]

            # 特徴8
            for k in range(1,9):
                n = ws.cell(row=n_base_row + row_offset + 5 + k,  column=n_base_column + col_offset + 5)
                n.value = selected_row[f'特徴{k}'].values[0]

            # 容器,肥効曲線の画像を貼り付ける
            # スクリプトのディレクトリを取得
            script_dir = os.path.dirname(os.path.abspath(__file__))
            # 絶対パスを生成
            img_path = os.path.join(script_dir, '容器', f'{fertilizer}.jpg')
#            img_path2 = os.path.join(script_dir, '肥効曲線', f'{fertilizer}.jpg')

            # ファイルの存在を確認
            if os.path.exists(img_path):
                
                # Pillowで画像を開く
                original_img = PILImage.open(img_path)

                # 画像のリサイズ
                new_size = (190, 257)  # 新しいサイズを指定
                resized_img = original_img.resize(new_size)
            
                # 一時的なファイルを作成
                with tempfile.NamedTemporaryFile(delete=False, suffix=".jpg") as tmp_file:
                    temp_path = tmp_file.name
                    resized_img.save(temp_path)

                # openpyxlのImageクラスでリサイズされた画像を読み込む
                img = OpenpyxlImage(temp_path)
                
                # Excelのセルに画像を貼り付ける位置を指定                       
                cell_address = ws.cell(row=n_base_row + row_offset + 2, column=n_base_column + col_offset).coordinate
                img.anchor = cell_address

                # 画像をシートに追加
                ws.add_image(img)

            
            else:
                # ファイルが存在しない場合は何もしない
                pass

            i = i + 1

#======================================

        row_start = 1
        row_end = 50
        col_start = 1  # A列は1
        col_end = n_base_column + col_offset + 11

        # 列を文字列に変換
        col_start_letter = get_column_letter(col_start)
        col_end_letter = get_column_letter(col_end)

        # print_areaを設定
        ws.print_area = f'{col_start_letter}{row_start}:{col_end_letter}{row_end}'

        # 13列ごとに垂直改ページを追加
        cols_per_page = 13  # 13列ごとに改ページ
        for col in range(cols_per_page, col_end + 1 , cols_per_page):
            ws.col_breaks.append(Break(id=col))

        # 印刷時の拡大・縮小を85%に設定
        ws.page_setup.scale = 74

#-==================================
        # 変更を保存する
        wb.save('kasei_tem_finish.xlsx')


    if selected_fertilizer_count_ekihi > 0:
        # ワークブックをロードする
        wb = openpyxl.load_workbook('ekihi_tem.xlsx')
        # ワークシートを選択する（シート名を指定する）
        ws = wb['液肥_テンプレ']

        # 必要数
        count_number_ekihi = selected_fertilizer_count_ekihi  ###ここがチェックされた数字となる。
        #テンプレートを作るところ。
        count_number_ekihi = count_number_ekihi - 1  # ここでチェックをつけられた分だけコピーすることになる。0からカウント


        for i in range(0, count_number_ekihi):
            row_count = 1
            col_count = 14
            col_offset = i * 13

            # コピー元の範囲（例: A1からM37）
            source_range = [[ws.cell(row=r, column=c) for c in range(1, 14)] for r in range(1, 37)]

            # コピー先の左上セル（例: N1）
            dest_start_cell = ws.cell(row=row_count, column=col_count + col_offset)

            def copy_cell(src_cell, dest_cell):
                dest_cell.value = src_cell.value
                if src_cell.has_style:
                    dest_cell.font = copy(src_cell.font)
                    dest_cell.border = copy(src_cell.border)
                    dest_cell.fill = copy(src_cell.fill)
                    dest_cell.number_format = copy(src_cell.number_format)
                    dest_cell.protection = copy(src_cell.protection)
                    dest_cell.alignment = copy(src_cell.alignment)

            # コピー元範囲の行数と列数を取得する
            row_count = len(source_range)
            col_count = len(source_range[0])

            # コピー元範囲をループしてコピー先にペーストする
            for i in range(row_count):
                for j in range(col_count):
                    src_cell = source_range[i][j]
                    dest_cell = ws.cell(row=dest_start_cell.row + i, column=dest_start_cell.column + j)
                    copy_cell(src_cell, dest_cell)

            # 指定された列幅にコピー元とコピー先の列幅を設定する
            #specified_widths = [1, 5.67, 8.42, 5.67, 4, 0.84, 8.08, 8.08, 8.08, 8.08, 8.08, 8.08, 8.08]
            specified_widths = [2.25, 1.5, 6.92, 8.42, 8.42, 8.42, 1, 8.42, 8.42, 8.42, 8.42, 8.42, 8.42]
            # コピー元の列幅を設定する
            for idx, width in enumerate(specified_widths, start=source_range[0][0].column):
                col_letter = openpyxl.utils.get_column_letter(idx)
                ws.column_dimensions[col_letter].width = width

            # コピー先の列幅を設定する
            for idx, width in enumerate(specified_widths, start=dest_start_cell.column):
                col_letter = openpyxl.utils.get_column_letter(idx)
                ws.column_dimensions[col_letter].width = width

    
        #ここからデータを入れるところ
        # 選択されたデータ数分入力する。リスト分0スタートなので、+1とする。
        mm = count_number_ekihi

        i = 0
        # 各肥料名についてループ
        for fertilizer in selected_fertilizer_ekihi:
            selected_row = df_ekihi[df_ekihi['肥料名称'] == fertilizer]
                
            col_offset = i * 13
            
            n_base_row = 4
            n_base_column = 2
            
            # 肥料名称
            n = ws.cell(row=n_base_row, column=n_base_column + col_offset)
            n.value = selected_row['肥料名称'].values[0]

            # 成分名1～5
            for k in range(1,6):
                n = ws.cell(row=n_base_row + 2, column=n_base_column + col_offset + 6 + k)
                n.value = selected_row[f'成分名{k}'].values[0]

            # 成分1～5
            for k in range(1,6):
                n = ws.cell(row=n_base_row + 3, column=n_base_column + col_offset + 6 + k)
                n.value = selected_row[f'成分{k}'].values[0]

            # その他
            for k in range(1,3):
                n = ws.cell(row=n_base_row + 4 + k,  column=n_base_column + col_offset + 7)
                n.value = selected_row[f'その他{k}'].values[0]

            # 容量
            n = ws.cell(row=n_base_row + 8,  column=n_base_column + col_offset + 8)
            n.value = selected_row['容量'].values[0]
            # 形状
            n = ws.cell(row=n_base_row + 9,  column=n_base_column + col_offset + 8)
            n.value = selected_row['形状'].values[0]
            # 液色
            n = ws.cell(row=n_base_row + 10,  column=n_base_column + col_offset + 8)
            n.value = selected_row['液色'].values[0]
            # 散布方法
            n = ws.cell(row=n_base_row + 11,  column=n_base_column + col_offset + 8)
            n.value = selected_row['散布方法'].values[0]
            # 特徴1～6
            for k in range(1,7):
                n = ws.cell(row=n_base_row + 13 + k,  column=n_base_column + col_offset)
                n.value = selected_row[f'特徴{k}'].values[0]
            # 使用方法1～6
            for k in range(1,7):
                n = ws.cell(row=n_base_row + 20 + k,  column=n_base_column + col_offset)
                n.value = selected_row[f'使用方法{k}'].values[0]
            
            # 容器,肥効曲線の画像を貼り付ける
            # スクリプトのディレクトリを取得
            script_dir = os.path.dirname(os.path.abspath(__file__))
            # 絶対パスを生成
            img_path = os.path.join(script_dir, '容器', f'{fertilizer}.jpg')
#            img_path2 = os.path.join(script_dir, '肥効曲線', f'{fertilizer}.jpg')

            # ファイルの存在を確認
            if os.path.exists(img_path):
                
                # Pillowで画像を開く
                original_img = PILImage.open(img_path)

                # 画像のリサイズ
                new_size = (170, 330)  # 新しいサイズを指定
                resized_img = original_img.resize(new_size)
            
                # 一時的なファイルを作成
                with tempfile.NamedTemporaryFile(delete=False, suffix=".jpg") as tmp_file:
                    temp_path = tmp_file.name
                    resized_img.save(temp_path)
    #                st.write(f"Image temporarily saved at {temp_path}")

    #
                # openpyxlのImageクラスでリサイズされた画像を読み込む
                img = OpenpyxlImage(temp_path)
                # Excelのセルに画像を貼り付ける位置を指定
                cell_address = ws.cell(row=n_base_row  + 3, column=n_base_column + col_offset ).coordinate
                img.anchor = cell_address

                # 画像をシートに追加
                ws.add_image(img)
                # 一時ファイルを削除
                #os.remove(temp_path) 
            
            else:
                # ファイルが存在しない場合は何もしない
                pass

            i = i + 1

        row_start = 1
        row_end = 37
        col_start = 1  # A列は1
        col_end = n_base_column + col_offset + 11

        # 列を文字列に変換
        col_start_letter = get_column_letter(col_start)
        col_end_letter = get_column_letter(col_end)

        # print_areaを設定
        ws.print_area = f'{col_start_letter}{row_start}:{col_end_letter}{row_end}'

        # 13列ごとに垂直改ページを追加
        cols_per_page = 13  # 13列ごとに改ページ
        for col in range(cols_per_page, col_end + 1 , cols_per_page):
            ws.col_breaks.append(Break(id=col))

        # 印刷時の拡大・縮小を85%に設定
        ws.page_setup.scale = 78
        # 変更を保存する
        wb.save('ekihi_tem_finish.xlsx')



    count_mokuji = selected_fertilizer_count
    count_ekihi_mokuji = selected_fertilizer_count_ekihi
    count_kasei_mokuji = selected_fertilizer_count_kasei 

    #コピーするの数を確認。
    if selected_fertilizer_count > 0:
        count_mokuji = ((count_mokuji - 1) // 3) 
    #    st.write(count_mokuji)
        count_mokuji = count_mokuji + 2

    if selected_fertilizer_count_ekihi > 0:
        count_ekihi_mokuji = ((count_ekihi_mokuji - 1) // 3) 
        count_ekihi_mokuji = count_ekihi_mokuji + 2
    
    if selected_fertilizer_count_kasei > 0:
        count_kasei_mokuji = ((count_kasei_mokuji - 1) // 3) 
        count_kasei_mokuji = count_kasei_mokuji + 2
    
    all_count = count_mokuji + count_ekihi_mokuji + count_kasei_mokuji

    # ワークブックをロードする
    wb = openpyxl.load_workbook('目次.xlsx')
    # ワークシートを選択する（シート名を指定する）
    ws = wb['目次']
    # 必要数
    count = ((all_count - 1) // 8)
    count += 1

    for i in range(0, count):
        row_count = 1
        col_count = 1
        col_offset = i * 5
        # コピー元の範囲（例: A1からE25）
        source_range = [[ws.cell(row=r, column=c) for c in range(1, 5)] for r in range(1, 25)]

        ## コピー先の左上セル（例: F1）
        dest_start_cell = ws.cell(row=1, column= col_count + col_offset)

        def copy_cell(src_cell, dest_cell):
            dest_cell.value = src_cell.value
            if src_cell.has_style:
                dest_cell.font = copy(src_cell.font)
                dest_cell.border = copy(src_cell.border)
                dest_cell.fill = copy(src_cell.fill)
                dest_cell.number_format = copy(src_cell.number_format)
                dest_cell.protection = copy(src_cell.protection)
                dest_cell.alignment = copy(src_cell.alignment)

        # コピー元範囲の行数と列数を取得する
        row_count = len(source_range)
        col_count = len(source_range[0])

        # コピー元範囲をループしてコピー先にペーストする
        for i in range(row_count):
            for j in range(col_count):
                src_cell = source_range[i][j]
                dest_cell = ws.cell(row=dest_start_cell.row + i, column=dest_start_cell.column + j)
                copy_cell(src_cell, dest_cell)

                # 指定された列幅にコピー元とコピー先の列幅を設定する
                specified_widths = [8.08, 8.08, 8.08, 8.08, 4.04, 8.08]

                # コピー元の列幅を設定する
                for idx, width in enumerate(specified_widths, start=source_range[0][0].column):
                    col_letter = openpyxl.utils.get_column_letter(idx)
                    ws.column_dimensions[col_letter].width = width

                # コピー先の列幅を設定する
                for idx, width in enumerate(specified_widths, start=dest_start_cell.column):
                    col_letter = openpyxl.utils.get_column_letter(idx)
                    ws.column_dimensions[col_letter].width = width

    # いらないところを消すのためのオフセット設定
    shita_offset = (all_count % 8) * 3
    migi_offset = (all_count // 8) * 5
    
  
    # いらない箇所を消す
    if shita_offset != 0:
        # A1:M44 の範囲のセルをループする
        for row in ws.iter_rows(min_row=1 + shita_offset, max_row=25, min_col=1 + migi_offset, max_col=5 + migi_offset):
            for cell in row:
            # セルの文字を消す
                cell.value = None

            # セルの罫線を消す
                cell.border = Border()

            # セルの背景色を消す (デフォルトは白)
                cell.fill = PatternFill(fill_type=None)

            # セルのフォントスタイルをデフォルトにリセット                   
                cell.font = Font()


    def name_insert(nam,cor):
        # RGB(91, 155, 213)を16進数で指定'5B9BD5'
        fill_color = PatternFill(start_color=cor, end_color=cor, fill_type='solid')

        for col in range(0, 5):  # 1列目(A)から5列目(E)
            ws.cell(row=start_row + row_offset + 1, column=start_col + col_offset + col).fill = fill_color
        
        name = ws.cell(row=start_row + row_offset + 1 , column=start_col + col_offset)

        name.value = nam
        # 文字色を白に設定
        white_font = Font(color="FFFFFF", size=16, bold=True)
        name.font = white_font
    
    #目次のデータ入力
    in_count = 0
    start_row = 1
    start_col = 1 
    selected_fertilizer_mo =  selected_fertilizer
    selected_fertilizer_kasei_mo =  selected_fertilizer_kasei
    selected_fertilizer_ekihi_mo =  selected_fertilizer_ekihi
    
    mo =  0
    kasei_mo =  0
    ekihi_mo =  0   
    
    page_number = 1
 
    for m in range(count_mokuji):   
        row_offset = (in_count % 8) *3
        col_offset = (in_count // 8) *5        

        if m == 0:
            #目次の題名を入れる。
            name_insert('BB肥料','5B9BD5')
            in_count += 1    
        else:    
            #銘柄名を入力していく。
            for i in range(0,3):    
                name = ws.cell(row=start_row + row_offset + i , column=start_col + col_offset)

                if selected_fertilizer_mo:
                    name.value = selected_fertilizer_mo.pop(0) 
                    name = ws.cell(row=start_row + row_offset + i , column=start_col + col_offset + 4)
                    
                    #mo +=1
                    #page_number = page_number + ((mo + 1) % 2) 
    
                    result = 1 if mo != 0 and mo % 2 == 0 else 0
                    
                    #st.write(result)
                    page_number = page_number + result                   
                    
                    name.value = page_number
                    # フォントを太文字に設定
                    bold_font = Font(bold=True, size=14)
                    # セルに太文字のフォントを適用
                    name.font = bold_font

                    mo +=1       
                    #name.value = page_number
                    # フォントを太文字に設定
                    #bold_font = Font(bold=True, size=16)
                    # セルに太文字のフォントを適用
                    #name.font = bold_font

            #page_number +=1
            in_count += 1
    
    page_number += 1

    
    for m in range(count_kasei_mokuji):   
        row_offset = (in_count % 8) *3
        col_offset = (in_count // 8) *5        

        #目次の題名を入れる。
        if m == 0:
            # RGB(237, 125, 49) を 16 進数に変換すると '#ED7D31'
            name_insert('化成','ED7D31')
            in_count += 1    
        else:    
            #銘柄名を入力していく。
            for i in range(0,3):    
                name = ws.cell(row=start_row + row_offset + i , column=start_col + col_offset)

                if selected_fertilizer_kasei_mo:
                    name.value = selected_fertilizer_kasei.pop(0)
                    name = ws.cell(row=start_row + row_offset + i , column=start_col + col_offset + 4)
                    
                    
                    result = 1 if kasei_mo != 0 and kasei_mo % 3 == 0 else 0
                    
                    page_number = page_number + result                   
                    
                    name.value = page_number
                    # フォントを太文字に設定
                    bold_font = Font(bold=True, size=14)
                    # セルに太文字のフォントを適用
                    name.font = bold_font

                    kasei_mo +=1       
            #page_number +=1
            in_count += 1

    page_number += 1

    for m in range(count_ekihi_mokuji):   
        row_offset = (in_count % 8) *3
        col_offset = (in_count // 8) *5        

        if m == 0:
            #目次の題名を入れる。
            name_insert('液肥','B5E6A2')
            in_count += 1    
        else:    
            #銘柄名を入力していく。
            for i in range(0,3):    
                name = ws.cell(row=start_row + row_offset + i , column=start_col + col_offset)

                if selected_fertilizer_ekihi_mo:
                    name.value = selected_fertilizer_ekihi.pop(0)
                    name = ws.cell(row=start_row + row_offset + i , column=start_col + col_offset + 4)
                    
                    #result = 1 if kasei_mo != 0 and kasei_mo % 1 == 0 else 0
                    #result += 1 
                    #page_number = page_number + result                   
                                    
                    name.value = page_number
                    #フォントを太文字に設定
                    bold_font = Font(bold=True, size=14)
                    ##セルに太文字のフォントを適用
                    name.font = bold_font
                    page_number += 1

            #page_number +=1
            in_count += 1
    
    page_number +=1

#================================================   
    row_start = 1
    row_end = 25
    col_start = 1  # A列は1
    col_end = start_col + col_offset + 9

    # 列を文字列に変換
    col_start_letter = get_column_letter(col_start)
    col_end_letter = get_column_letter(col_end)

    # print_areaを設定
    ws.print_area = f'{col_start_letter}{row_start}:{col_end_letter}{row_end}'

    # 13列ごとに垂直改ページを追加
    cols_per_page = 10  # 13列ごとに改ページ
    for col in range(cols_per_page, col_end + 1 , cols_per_page):
        ws.col_breaks.append(Break(id=col))

    # 印刷時の拡大・縮小を85%に設定
    ws.page_setup.scale = 105

#===========================================

    # 保存する場合
    wb.save("目次_finish.xlsx")

    st.success('🔥 🔥 セットアップ完了しました🔥 🔥 ')

# 3つのカラムを作成
col4, col5, col6 = st.columns(3)

with col4:
    # Excelファイルを読み込む
    with open('bb_tem_finish.xlsx', 'rb') as file:  # ここでファイルを開きます
        excel_data = file.read()  # インデントされていることを確認
# ダウンロードボタンの作成
    st.download_button(
        label="Download Excel File＜BB＞",  # ボタンのラベル
        data=excel_data,  # ダウンロードするデータ
        file_name='bb_tem_finish.xlsx',  # ダウンロード時のファイル名
        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'  # MIMEタイプを指定
    )

with col5:
    # Excelファイルを読み込む
    with open('kasei_tem_finish.xlsx', 'rb') as file:
        excel_data_ekihi = file.read()
# ダウンロードボタンの作成
    st.download_button(
        label="Download Excel File＜化成＞",  # ボタンのラベル
        data=excel_data_ekihi,  # ダウンロードするデータ
        file_name='kasei_tem_finish.xlsx',  # ダウンロード時のファイル名
        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'  # MIMEタイプを指定
    )
 
with col6:
    # Excelファイルを読み込む
    with open('ekihi_tem_finish.xlsx', 'rb') as file:
        excel_data_ekihi = file.read()
# ダウンロードボタンの作成
    st.download_button(
        label="Download Excel File＜液肥＞",  # ボタンのラベル
        data=excel_data_ekihi,  # ダウンロードするデータ
        file_name='ekihi_tem_finish.xlsx',  # ダウンロード時のファイル名
        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'  # MIMEタイプを指定
    )

# 3つのカラムを作成
col7, col8, col9 = st.columns(3)

with col7:

    with open('目次_finish.xlsx', 'rb') as file:
            mokuji_ekihi = file.read()

    st.download_button(
            label="Download Excel File＜目次＞",  # ボタンのラベル
            data=mokuji_ekihi,  # ダウンロードするデータ
            file_name='目次_finish.xlsx',  # ダウンロード時のファイル名
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'  # MIMEタイプを指定
        )

with col8:

    with open('肥料パンフレット_表紙.xlsx', 'rb') as file:
            panfu = file.read()

    st.download_button(
            label="Download Excel File＜表紙＞",  # ボタンのラベル
            data=panfu,  # ダウンロードするデータ
            file_name='肥料パンフレット_表紙.xlsx',  # ダウンロード時のファイル名
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'  # MIMEタイプを指定
        )

#import win32com.client as win32
##import win32api

# def save_excel_as_pdf(excel_path, pdf_path, row_start, row_end, col_start, col_end):
#     excel = win32.Dispatch("Excel.Application")
#     excel.Visible = False
#     wb = excel.Workbooks.Open(excel_path)
#     ws = wb.Worksheets[0]
    
#     # 印刷範囲やレイアウトの設定
#     #ws.PageSetup.PrintArea = "A1:G100"  # 印刷範囲を設定
    
#     # 行と列を数値で動的に印刷範囲を設定
#     start_cell = ws.Cells(row_start, col_start).Address  # 開始セル
#     end_cell = ws.Cells(row_end, col_end).Address  # 終了セル
#     ws.PageSetup.PrintArea = f"{start_cell}:{end_cell}"  # 印刷範囲を設定
    

#     #ws.PageSetup.FitToPagesWide = 1  # 横1ページに収める
#     #ws.PageSetup.FitToPagesTall = False  # 縦のページ数制限なし
#     #ws.PageSetup.PrintTitleRows = "$1:$1"  # 1行目をタイトル行として繰り返し
#     ##ws.PageSetup.Zoom = False  # 拡大・縮小はしない
    
#     # PDFとして保存
#     ws.ExportAsFixedFormat(0, pdf_path)
#     wb.Close(SaveChanges=False)
#     excel.Quit()

# # 使用例
# #ddsave_excel_as_pdf("目次_finish.xlsx", "finish.pdf", 1, 25, 1, 10)  # A25:G100

# #================
# with open('目次_finish.xlsx', 'rb') as file:
#         mokuji = file.read()

# st.download_button(
#         label="Download Excel File＜目次＞_PDF",  # ボタンのラベル
#         data=mokuji,  # ダウンロードするデータ
#         file_name='目次_finish.xlsx',  # ダウンロード時のファイル名
#         mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'  # MIMEタイプを指定
#     )


